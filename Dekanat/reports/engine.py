"""Тонкий рушій рендерингу документів за шаблоном (DK-27, DK-29).

`render_docx` (поверх `docxtpl`) та `render_xlsx` (поверх `xlsxtpl`) — обидва
повертають `BytesIO` і ділять спільні Jinja-фільтри (гроші, дати, оцінки).
Знає лише про директорію шаблонів і фільтри — stateless. Бізнес-логіку та схему
даних тримають класи-звіти у `reports/base.py`, `reports/rating.py`,
`reports/xlsx_reports.py`.

Шар навмисно не залежить від Reflex: повертаємо `BytesIO`, а вже state-шар
загортає байти у `rx.download(...)`.
"""

import math
from io import BytesIO
from pathlib import Path
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Callable, Optional

from openpyxl.workbook import Workbook

from docxtpl import DocxTemplate
from jinja2 import Environment, StrictUndefined
from xlsxtpl.writerx import BookWriter
from openpyxl import load_workbook as _load_xlsx
from openpyxl.utils import get_column_letter as _col_letter

# Усі шаблони лежать поряд, у `reports/templates/` (.docx та .xlsx).
TEMPLATES_DIR = Path(__file__).parent / "templates"


# Українські місяці в родовому відмінку (для «23 липня 2025 р.»).
# Словник замість locale: не залежить від системної локалі, потокобезпечний
# і одразу дає правильний відмінок (locale на %B віддає називний «липень»).
_UA_MONTHS_GEN = {
    1: "січня", 2: "лютого", 3: "березня", 4: "квітня",
    5: "травня", 6: "червня", 7: "липня", 8: "серпня",
    9: "вересня", 10: "жовтня", 11: "листопада", 12: "грудня",
}


# Нерозривний пробіл як розділювач тисяч — щоб "1 234 567" не переносилось
# на різні рядки всередині документа.
_NBSP = "\u00a0"


def _fmt_money(value: Any) -> str:
    """1234567.5 → "1 234 567,50" (нерозривні пробіли, кома як десятковий
    розділювач). Порожнє/None → "" (хай шаблон сам вирішує)."""
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):,.2f}".replace(",", _NBSP).replace(".", ",")
    except (TypeError, ValueError):
        return str(value)


def _coerce_date(value: Any) -> date | None:
    """Приймає date/datetime/ISO-рядок ("2026-06-27" чи "2026-06-27T..."),
    повертає date або None, якщо розпарсити не вдалося."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value:
        try:
            return date.fromisoformat(value[:10])
        except ValueError:
            return None
    return None


def _fmt_date(value: Any, fmt: str = "%d.%m.%Y") -> str:
    """date/datetime/ISO-рядок → "27.06.2026". Невідоме → вихідне значення як є."""
    d = _coerce_date(value)
    if d is not None:
        return d.strftime(fmt)
    return "" if value is None else str(value)


def _fmt_datetime(value: Any, fmt: str = "%d.%m.%Y %H:%M") -> str:
    """datetime/ISO-рядок → "27.06.2026 14:30"."""
    if isinstance(value, datetime):
        return value.strftime(fmt)
    if isinstance(value, str) and value:
        try:
            return datetime.fromisoformat(value).strftime(fmt)
        except ValueError:
            return _fmt_date(value, "%d.%m.%Y")
    return "" if value is None else str(value)


def _date_uk(value: Any) -> str:
    """date(2025, 7, 23) → «23» липня 2025 р. (родовий відмінок місяця).
    Невідоме/None → "" (StrictUndefined зловить відсутність змінної окремо)."""
    d = _coerce_date(value)
    if d is None:
        return ""
    return f"«{d.day}» {_UA_MONTHS_GEN[d.month]} {d.year} р."


def _grade(value: Any) -> str:
    """Оцінка/сума в українському форматі: Decimal('188') → '188,00'.
    Порожнє/None → "" (порожня клітинка в таблиці)."""
    if value is None or value == "":
        return ""
    try:
        return f"{Decimal(value):.2f}".replace(".", ",")
    except (TypeError, ValueError, ArithmeticError):
        return str(value)


def _jinja_env() -> Environment:
    """Jinja-середовище для шаблонів.

    `StrictUndefined` — fail-fast: якщо шаблон посилається на змінну, якої немає
    у context, рендер впаде з помилкою замість мовчазної підстановки порожнечі
    (для офіційних документів тиха порожнеча небезпечна).
    """
    env = Environment(undefined=StrictUndefined)
    env.filters["money"] = _fmt_money
    env.filters["date"] = _fmt_date
    env.filters["datetime"] = _fmt_datetime
    env.filters["date_uk"] = _date_uk
    env.filters["grade"] = _grade
    return env


def render_docx(template_name: str, context: dict) -> BytesIO:
    """Відрендерити шаблон `template_name` з `context` → `BytesIO` (.docx).

    Кидає `FileNotFoundError`, якщо шаблону немає, і пробрасує помилки Jinja
    (зокрема `UndefinedError` від `StrictUndefined`) — fail-fast.
    """
    path = TEMPLATES_DIR / template_name
    if not path.exists():
        raise FileNotFoundError(f"Шаблон не знайдено: {template_name}")
    try:
        doc = DocxTemplate(path)
        doc.render(context, jinja_env=_jinja_env())
        buffer = BytesIO()
        doc.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        print(f"[reports.engine][render_docx][ERROR] {template_name}: {e}")
        raise


# Параметри авто-підбору висоти рядків (для клітинок з переносом слів).
_DEFAULT_COL_WIDTH = 8.43   # дефолтна ширина стовпця openpyxl, у «символах»
_LINE_HEIGHT_PT = 15.0      # висота одного рядка тексту у пунктах (шрифт ~11pt)


def _autofit_row_heights(ws) -> None:
    """Підганяє висоту рядків під вміст клітинок із `wrap_text` (включно з
    об'єднаними) — Excel сам не рахує висоту для merged-клітинок. Оцінка: явні
    переноси `\\n` + орієнтовне перенесення за шириною стовпця(ів)."""
    # Однорядкові горизонтальні merge: (рядок, перший стовпець) -> останній стовпець.
    h_merges = {
        (m.min_row, m.min_col): m.max_col
        for m in ws.merged_cells.ranges
        if m.min_row == m.max_row
    }
    for row in ws.iter_rows():
        max_lines = 1
        has_wrap = False
        for cell in row:
            val = cell.value
            if not isinstance(val, str) or not val:
                continue
            al = cell.alignment
            if al is None or not al.wrap_text:
                continue
            has_wrap = True
            col_max = h_merges.get((cell.row, cell.column), cell.column)
            width = 0.0
            for ci in range(cell.column, col_max + 1):
                w = ws.column_dimensions[_col_letter(ci)].width
                width += w if w else _DEFAULT_COL_WIDTH
            chars = max(1, int(width))
            lines = 0
            for seg in val.split("\n"):
                lines += max(1, math.ceil(len(seg.rstrip("\r")) / chars))
            max_lines = max(max_lines, lines)
        if has_wrap and max_lines > 1:
            needed = max_lines * _LINE_HEIGHT_PT
            rd = ws.row_dimensions[row[0].row]
            if rd.height is None or rd.height < needed:
                rd.height = needed


def render_xlsx(
    template_name: str,
    context: dict,
    post_process: Optional[Callable[[Workbook], None]] = None,
) -> BytesIO:
    """Відрендерити xlsx-шаблон `template_name` з `context` → `BytesIO` (.xlsx).

    Поверх `xlsxtpl` (BookWriter/xltpl): Jinja2 прямо в клітинках, таблиця росте
    за даними, підвал зі своїми merge/стилями зсувається вниз сам. Кожен шаблон —
    одна сторінка, тож `render_book` отримує список з одного payload'а.

    Ті самі Jinja-фільтри, що й для docx (`date_uk`, `grade`, `money`, ...), —
    реєструються через `add_filter`. Кидає `FileNotFoundError`, якщо шаблону немає.

    Після рендеру висота рядків підганяється під вміст клітинок із переносом слів
    (`_autofit_row_heights`) — окремим проходом через openpyxl, бо merged-клітинки
    Excel сам не авто-підганяє. `post_process` (DK-66) — опційний хук звіту для
    додаткової стилізації (напр. жовта заливка непорожньої клітинки), викликається
    тим самим openpyxl-проходом, до фінального збереження.

    Особливість xlsx-рушія: директиви циклу (`for`/`endfor`) живуть у коментарях
    клітинок шаблону, а не у значеннях (закладено у самих шаблонах, DK-29).
    """
    path = TEMPLATES_DIR / template_name
    if not path.exists():
        raise FileNotFoundError(f"Шаблон не знайдено: {template_name}")
    try:
        writer = BookWriter(str(path))
        env = _jinja_env()
        for name, fn in env.filters.items():
            # Переносимо лише наші кастомні фільтри (вбудовані xltpl має свої).
            if name in ("date_uk", "grade", "money", "date", "datetime"):
                writer.add_filter(name, fn)
        writer.render_book([context])

        rendered = BytesIO()
        writer.save(rendered)
        rendered.seek(0)

        # Пост-обробка: авто-висота рядків під wrap_text-клітинки.
        wb = _load_xlsx(rendered)
        for ws in wb.worksheets:
            _autofit_row_heights(ws)
        if post_process is not None:
            post_process(wb)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    except Exception as e:
        print(f"[reports.engine][render_xlsx][ERROR] {template_name}: {e}")
        raise
