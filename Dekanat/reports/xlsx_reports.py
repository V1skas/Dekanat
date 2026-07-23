"""Екзаменаційні відомості у XLSX (DK-29).

Три звіти на групу абітурієнтів:
  * VidomistReport      — відомість обліку вступного випробування (шапка + таблиця);
  * VykladachamReport   — список для викладачів;
  * TelefonyReport      — список з телефонами (абітурієнта та родичів).

Дані в контекст рендеру передаються об'єктами (ExamApplicant), а не словниками —
у шаблоні доступ через `a.name` / `a.phone` / `a.relatives`.
"""

from __future__ import annotations

from copy import copy
from datetime import date
from io import BytesIO
from typing import ClassVar

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from Dekanat.reports.engine import TEMPLATES_DIR, _autofit_row_heights, render_xlsx


class ExamApplicant(BaseModel):
    name: str                 # ПІБ
    phone: str = ""           # телефон абітурієнта (з картки)
    relatives: str = ""       # телефони родичів з ПІБ (для листа телефонів)
    grades: dict[int, str] = Field(default_factory=dict)  # id предмета -> бал (DK-62)


class SubjectColumn(BaseModel):
    """Колонка-предмет у документі «Викладачам» (DK-62)."""

    id: int
    title: str


class XlsxReport(BaseModel):
    """Базовий xlsx-звіт: зв'язок Pydantic-схеми з конкретним шаблоном.

    Валідація відбувається при створенні об'єкта (fail-fast). `context()`
    за замовчуванням віддає всі поля, але з applicants як об'єктами."""

    template_name: ClassVar[str]
    file_basename: ClassVar[str] = "report"
    sheet_name: ClassVar[str] = ""  # назва листа у вихідному файлі

    applicants: list[ExamApplicant]

    def context(self) -> dict:
        return {"applicants": self.applicants, "sheet_name": self.sheet_name}

    def post_process(self, wb: Workbook) -> None:
        """Хук пост-обробки готової книги (DK-66) — за замовчуванням no-op."""
        return None

    @property
    def filename(self) -> str:
        return f"{self.file_basename}.xlsx"

    def render(self) -> BytesIO:
        return render_xlsx(self.template_name, self.context(), post_process=self.post_process)

    def render_bytes(self) -> bytes:
        return self.render().getvalue()


_UNKNOWN_SPECIALTY_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


class VidomistReport(XlsxReport):
    template_name: ClassVar[str] = "vidomist_template.xlsx"
    file_basename: ClassVar[str] = "vidomist"
    sheet_name: ClassVar[str] = "Відомість"

    specialty: str            # «D1 Інженерія ...» — визначається за тегом групи (DK-66)
    opp: str = ""             # освітньо-професійна програма спеціальності (DK-44)
    specialty_unknown: bool = False  # тег групи не дав однозначної спеціальності (DK-66)
    subject: str = ""         # форма контролю / предмет випробування
    report_date: date         # дата формування (Pydantic розпарсить ISO-рядок)
    number: str = ""          # № відомості — лишається порожнім (від руки)

    def context(self) -> dict:
        # Об'єкти, а не дамп: applicants як ExamApplicant; шапка — скалярами.
        return {
            "specialty": self.specialty,
            "opp": self.opp,
            "subject": self.subject,
            "report_date": self.report_date,
            "number": self.number,
            "applicants": self.applicants,
            "sheet_name": self.sheet_name,
        }

    def post_process(self, wb: Workbook) -> None:
        """Якщо спеціальність за тегом групи визначити не вдалось (DK-66) —
        підсвічуємо порожні клітинки спеціальності/ОПП жовтим, щоб оператор
        помітив і заповнив вручну."""
        if not self.specialty_unknown:
            return
        ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active
        if ws is None:
            return
        ws["D3"].fill = _UNKNOWN_SPECIALTY_FILL
        ws["D4"].fill = _UNKNOWN_SPECIALTY_FILL


def _copy_style(src_cell, dst_cell) -> None:
    dst_cell.font = copy(src_cell.font)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.border = copy(src_cell.border)


def _build_vykladacham(
    subjects: list[SubjectColumn],
    applicants: list[ExamApplicant],
    sheet_name: str,
) -> BytesIO:
    """Список для викладачів (DK-62): колонки-предмети — динамічні (кількість і
    склад обирає користувач), тому цей звіт будується напряму через openpyxl, а
    не xltpl (у xltpl немає директиви колонкового циклу — лише рядковий).

    Стилі клітинок знімаються з `spysok_vykladacham_template.xlsx`, щоб вигляд
    лишався узгодженим з рештою відомостей. Розкладка колонок: № · Прізвище
    (B:D) · Для заміток (E:I) · по одній колонці на кожен обраний предмет ·
    «Уч закладі» — завжди крайня права.
    """
    src_wb = load_workbook(TEMPLATES_DIR / "spysok_vykladacham_template.xlsx")
    src_ws = src_wb.active
    assert src_ws is not None

    NUM_COL = 1
    NAME_COL_START, NAME_COL_END = 2, 4
    NOTES_COL_START, NOTES_COL_END = 5, 9
    SUBJECTS_COL_START = 10
    tail_col = SUBJECTS_COL_START + len(subjects)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name

    header_row_first, header_row_last = 2, 4
    data_row_start = 5

    def header_cell(col: int, value):
        c = ws.cell(row=header_row_first, column=col, value=value)
        c.alignment = copy(src_ws["J2"].alignment)
        c.font = copy(src_ws["J2"].font)
        c.border = copy(src_ws["J2"].border)
        ws.merge_cells(start_row=header_row_first, start_column=col, end_row=header_row_last, end_column=col)
        for r in range(header_row_first, header_row_last + 1):
            ws.cell(row=r, column=col).border = copy(src_ws["J2"].border)
        return c

    # № з/п
    _copy_style(src_ws["A2"], ws.cell(row=header_row_first, column=NUM_COL, value=src_ws["A2"].value))
    ws.merge_cells(start_row=header_row_first, start_column=NUM_COL, end_row=header_row_last, end_column=NUM_COL)
    ws.column_dimensions[get_column_letter(NUM_COL)].width = src_ws.column_dimensions["A"].width

    # Прізвище, ім'я вступника
    _copy_style(src_ws["B2"], ws.cell(row=header_row_first, column=NAME_COL_START, value=src_ws["B2"].value))
    ws.merge_cells(
        start_row=header_row_first, start_column=NAME_COL_START,
        end_row=header_row_last, end_column=NAME_COL_END,
    )
    ws.column_dimensions[get_column_letter(NAME_COL_END)].width = src_ws.column_dimensions["D"].width

    # Для заміток
    _copy_style(src_ws["E2"], ws.cell(row=header_row_first, column=NOTES_COL_START, value=src_ws["E2"].value))
    ws.merge_cells(
        start_row=header_row_first, start_column=NOTES_COL_START,
        end_row=header_row_last, end_column=NOTES_COL_END,
    )
    ws.column_dimensions[get_column_letter(NOTES_COL_START)].width = src_ws.column_dimensions["E"].width

    # Колонки-предмети
    for offset, subject in enumerate(subjects):
        col = SUBJECTS_COL_START + offset
        header_cell(col, subject.title)
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Уч закладі — крайня права
    _copy_style(src_ws["K2"], ws.cell(row=header_row_first, column=tail_col, value=src_ws["K2"].value))
    ws.merge_cells(start_row=header_row_first, start_column=tail_col, end_row=header_row_last, end_column=tail_col)

    # Дані
    for idx, applicant in enumerate(applicants):
        row = data_row_start + idx

        num_cell = ws.cell(row=row, column=NUM_COL, value=idx + 1)
        _copy_style(src_ws["A5"], num_cell)

        name_cell = ws.cell(row=row, column=NAME_COL_START, value=applicant.name)
        _copy_style(src_ws["B5"], name_cell)
        ws.merge_cells(start_row=row, start_column=NAME_COL_START, end_row=row, end_column=NAME_COL_END)
        for c in range(NAME_COL_START, NAME_COL_END + 1):
            ws.cell(row=row, column=c).border = copy(src_ws["B5"].border)

        notes_cell = ws.cell(row=row, column=NOTES_COL_START)
        _copy_style(src_ws["B5"], notes_cell)
        ws.merge_cells(start_row=row, start_column=NOTES_COL_START, end_row=row, end_column=NOTES_COL_END)

        for offset, subject in enumerate(subjects):
            col = SUBJECTS_COL_START + offset
            grade_cell = ws.cell(row=row, column=col, value=applicant.grades.get(subject.id, ""))
            _copy_style(src_ws["K5"], grade_cell)

        tail_cell = ws.cell(row=row, column=tail_col)
        _copy_style(src_ws["K5"], tail_cell)

    last_row = data_row_start + len(applicants) - 1 if applicants else header_row_last
    for r in range(1, last_row + 1):
        ws.row_dimensions[r].height = 14.25

    _autofit_row_heights(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


class VykladachamReport(XlsxReport):
    template_name: ClassVar[str] = "spysok_vykladacham_template.xlsx"
    file_basename: ClassVar[str] = "spysok_vykladacham"
    sheet_name: ClassVar[str] = "Список викладачам"

    subjects: list[SubjectColumn] = Field(default_factory=list)

    def render(self) -> BytesIO:
        return _build_vykladacham(self.subjects, self.applicants, self.sheet_name)


class TelefonyReport(XlsxReport):
    template_name: ClassVar[str] = "spysok_telefony_template.xlsx"
    file_basename: ClassVar[str] = "spysok_telefony"
    sheet_name: ClassVar[str] = "Список телефони"
